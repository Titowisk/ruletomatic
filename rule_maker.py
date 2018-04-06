# Python Path:
# "python.pythonPath": "C:/Users/vitor.rabelo/virtual_enviroments/ruletomatic/Scripts/python.exe"

import argparse
import re
from collections import namedtuple
from itertools import groupby
from os import mkdir
from datetime import datetime

from openpyxl import load_workbook

def main():

    # rule_maker.py menu
    # https://docs.python.org/3.3/library/argparse.html?highlight=arg
    # TODO too many args :/
    parser = argparse.ArgumentParser(
        description='Read a excel file, process its rules and format the result for input in Talend TMap Expressions.'
    )
    group1 = parser.add_argument_group('Excel row and columns configuration.')
    group1.add_argument(
        'row_heading', type=int,
        help='numero da linha onde esta o cabecalho das regras'
    )
    group1.add_argument(
        'last_data_row', type=int,
        help='numero da ultima linha que ainda contem valores'
    )
    group1.add_argument(
        'first_column', type=int, metavar='first_col',
        help='number of the first column'
    )
    group1.add_argument(
        'last_column', type=int, metavar='last_col',
        help='number of the last column'
    )
    parser.add_argument(
        'project_title', metavar='proj_title',
        help='the name of the current project to use as a version control inside each file created'
    )
    parser.add_argument(
        'directory_address', metavar='dir_addr',
        help="the name of the directory to use (creates it if it doesn't exist)",
        
    )
    
    args = parser.parse_args()
    print(args.row_heading)
    if args.directory_address:
        directory_address = args.directory_address
    else: # TODO Can I do this as a Function default (function name: write_to_file) ?
        directory_address = "regras_" + args.project_title + "/"

    dict_rules = dataToDict(
        row_heading=args.row_heading,
        last_data_row=args.last_data_row,
        first_column=args.first_column,
        last_column=args.last_column
    )

    rules_obj_list = applie_rules(dict_rules)

    for rule in rules_obj_list:
        print(rule.name)
        print("-"*len(rule.name))
        fully_formated_rule = format_to_talend(rule)
        # print(fully_formated_rule)
        write_to_file(rule.name, fully_formated_rule, directory_address=directory_address, project_title=args.project_title)


# -------------------------------------------------------------------------------------------------------------------

# This module gathers data from the excel rules file.
def dataToDict(row_heading, last_data_row, first_column, last_column):

    excel_rules_file_path = 'excel/regras_msp_ferreirav12.xlsx'
    wb = load_workbook(excel_rules_file_path) # get the workbook
    ws = wb.active # get the first worksheet from the file

    dict_rules = {} # a dictionary with {key:value} -> {<string>, <list>}

    for col in ws.iter_cols(min_row=row_heading, max_row=last_data_row, min_col=first_column, max_col=last_column):
        for i, cell in enumerate(col):
            if i == 0:
                key = str(cell.value)
                dict_rules.setdefault(key, [])
            else:
                dict_rules[key].append(str(cell.value))
    return dict_rules

# def excel_configuration(worksheet):

#     class breakThroughLoops(Exception): pass
    
#     # find the the top and left limits
#     try:
#         for row in worksheet.rows:
#             for cell in row:
#                 m = re.search(r'\D_\d+', cell.value)
#                 if m:
#                     location = cell.colum + cell.row  # example 'B6'
#                     row_heading = int(cell.row) - 1 # 6-1 =5
#                     first_column = int(cell.col_idx)  # 2
#                     raise breakThroughLoops
#     except breakThroughLoops:
#         pass
    
#     # find the bottom limit
#     try:
#         for col in worksheet.iter_cols(min_row=cell.row, min_col=cell.col_idx, max_col=cell.col_idx):
#             for cell in col:
#                 n = re.search(r'\D_\d+', cell.value)
#                 if not n:
#                     last_data_row = int(cell.col_idx) -1
#                     raise breakThroughLoops
#     except breakThroughLoops:
#         pass

#     # find the right limit

#     return row_heading, last_data_row, first_column, last_column


# -------------------------------------------------------------------------------------------------------------------

# This module applies the rules creating different sets for each rule

def applie_rules(dict_rules): # what name should I use?

    Rules = namedtuple('Rules', ['name', 'category', 'grouped_by_categories'])
    rules_values = dict_rules.values() # I iterate over it
    rules_keys = list(dict_rules.keys()) # I call each element to name each Rule
    rules_obj_list = []
    
    for i, column in enumerate(rules_values): # [(0, ["E_01", "E_02", "E_03" ...]), (1, [0, 0, 0, 1, 'N/A' ...]), (2, [1, 1, 1, 0, 'N/A'])]
        if i == 0: # line_of_codes must be the first column
            list_of_line_codes = column
        elif ('INFORMADO NA BASE' in column):
            continue
        else:
            line_codes_and_rules = sorted(zip(list_of_line_codes, column), key= lambda x: str(x[1])) 
            # print(list(line_codes_and_rules)) # ("E_01", 0), ("E_02", 0)...
            categories = set()
            grouped_by_categories = {}
            for key, group in groupby(line_codes_and_rules, key= lambda x: x[1]): # the key is always the second element of the tuple
                # print("key: {0}, group: {1}".format(key, group))
                # prints:
                # -----------------------------------------------------------------
                # Grouped Rules:
                # key: 0, group: <itertools._grouper object at 0x000002D456D9E5C0>
                # key: 1, group: <itertools._grouper object at 0x000002D456D9EDA0>
                # key: N/A, group: <itertools._grouper object at 0x000002D456D9EFD0>
                # key: 0, group: <itertools._grouper object at 0x000002D455B2DCC0>
                # key: 1, group: <itertools._grouper object at 0x000002D456091D68>
                # key: N/A, group: <itertools._grouper object at 0x000002D455B2DCC0>
                # -----------------------------------------------------------------
                categories.add(key)
                grouped_by_categories[key] = [ pair[0] for pair in group]

            rules_obj_list.append(Rules(rules_keys[i], categories, grouped_by_categories))

    return rules_obj_list # return a list of named tuples

# -------------------------------------------------------------------------------------------------------------------

# This module format the expressions to use in Talend

def format_to_talend(rule_obj): # rule_obj.name, rule_obj.category, rule_obj.grouped_by_categories

    fully_formated_rule = ""
        
    for category in sorted(rule_obj.category): #
        initial_condition = "row1.LINHA_APURACAO != null && \n"
        list_of_codes = rule_obj.grouped_by_categories

        if(len(list_of_codes[category]) == 1):
            #do something
            category_condition = "row1.LINHA_APURACAO.equalsIgnoreCase(\"{code}\") ? \"{rule}\" : \n".format(
                code=list_of_codes[category][0], rule=category
            )
            formated_category =  initial_condition + category_condition
            fully_formated_rule += formated_category
        else:
            #do other thing
            for i, code in enumerate(sorted(list_of_codes[category])):  #
                if(i == 0):
                    category_condition = "(" + "row1.LINHA_APURACAO.equalsIgnoreCase(\"{code}\")".format(code=code)
                else: # len(list_of_codes) = 4  [0, 1, 2, 3]
                    category_condition += " || row1.LINHA_APURACAO.equalsIgnoreCase(\"{code}\")".format(code=code)
                    if( i % 2 !=0):
                        category_condition += '\n'
            
            category_condition = category_condition.rstrip('\n') + ")" + " ? \"{rule}\" : \n".format(rule=category)
            formated_category =  initial_condition + category_condition
            fully_formated_rule += formated_category

    fully_formated_rule += "\"\""
    return fully_formated_rule

# -------------------------------------------------------------------------------------------------------------------

# This module write each fully_formated_rule into separeted .txt files

def write_to_file(rule_name, fully_formated_rule, directory_address, project_title):

    try:
        mkdir(directory_address)
    except FileExistsError:
        pass

    file_address = directory_address + '/' + '{}.txt'.format(rule_name)
    try:
        # if file exists
        with open(file_address, "r+") as f:
            file_data = f.read()

            # Adds version and hour of modification
            new_title = version_control(file_data)

            # Checks if the new expression is equal to previous expression
            if check_changes(file_data, fully_formated_rule):
                changed = ' | (no changes detected from previous expression)'
                new_title += changed
            full_text = '\n' + new_title + '\n' + fully_formated_rule
            f.seek(0,0)
            f.write(full_text + '\n' + file_data)
    except FileNotFoundError:
        # if file does not exist
        with open(file_address, "w+") as f:
            full_text = '\n' + project_title + ' 1' + '\n' + fully_formated_rule
            f.write(full_text)
    finally:
        f.close()
        print("Formated rule for talend successfully created in {}".format(file_address))

def version_control(file_data):
    time = datetime.now()
    modified = '{day:02}/{month:02} - {hour:02}:{minute:02}:{second:02}'.format(
        day=time.day, month=time.month, 
        hour=time.hour, minute=time.minute, second=time.second
        )
    end = file_data.find('\n', 2)
    title_list = file_data[:end].split(" ")
    new_version = int(title_list[1]) + 1
    new_title = title_list[0].lstrip('\n') + ' ' + str(new_version) + ' | modified: ' + modified
    return new_title

def check_changes(file_data, fully_formated_rule):

    start = file_data.find('row1')
    end = file_data.find('\"\"', start)
    previous_fully_formated_rule = file_data[start: end + 2]   

    return sorted(fully_formated_rule) == sorted(previous_fully_formated_rule)



main()